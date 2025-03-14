import streamlit as st
import pandas as pd
import numpy as np
import io
import base64

st.set_page_config(page_title="Excel Data Cleaner for IPP", layout="wide")

st.title("Excel Data Cleaner for Retool")
st.write("Upload your Excel file to clean and format it according to Retool requirements.")

# File uploader
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the Excel file
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        st.success("File successfully loaded!")
        
        # Display the original data
        st.subheader("Original Data")
        st.dataframe(df)
        
        # Process the data
        if st.button("Process Data"):
            st.subheader("Processing...")
            
            # Make a copy to avoid modifying the original
            processed_df = df.copy()
            
            # Step 1: Calculate total quantity sold per SKU (SUMIF equivalent)
            # Assuming Column A contains SKU and Column E contains quantity
            st.write("1. Calculating total quantity sold per SKU...")
            
            # Get column names for clarity
            sku_column = processed_df.columns[0]  # Assuming A is the first column
            quantity_column = processed_df.columns[4]  # Assuming E is the fifth column
            
            # Create a new column I for the SUMIF equivalent
            processed_df['Total_Quantity_Per_SKU'] = processed_df.apply(
                lambda row: processed_df[processed_df[sku_column] == row[sku_column]][quantity_column].sum(), 
                axis=1
            )
            
            # Step 2: Copy this column as VALUES ONLY to column J
            st.write("2. Copying totals as values only...")
            processed_df['Total_Quantity_Values'] = processed_df['Total_Quantity_Per_SKU']
            
            # Step 3: Store the total quantity (sum of column E)
            total_quantity = processed_df[quantity_column].sum()
            st.write(f"3. Total quantity from Column E: {total_quantity}")
            
            # Step 4: Clear rows past the last product line
            # This is a bit tricky without knowing the exact format of "Totals & Applied Filters text"
            # For now, we'll assume all valid data rows don't have NaN in critical columns
            st.write("4. Clearing rows past the last product line...")
            processed_df = processed_df.dropna(subset=[sku_column], how='any')
            
            # Step 5: Remove columns D, E, and I
            st.write("5. Removing unnecessary columns...")
            # We'll keep track of which columns to drop
            cols_to_drop = []
            
            # Check if we have enough columns to drop D and E
            if len(processed_df.columns) > 4:
                cols_to_drop.append(processed_df.columns[3])  # Column D (index 3)
            if len(processed_df.columns) > 4:
                cols_to_drop.append(processed_df.columns[4])  # Column E (index 4)
            
            # Drop the columns if we identified them
            if cols_to_drop:
                processed_df = processed_df.drop(columns=cols_to_drop)
            
            # We don't drop column I since we created it and it's now named 'Total_Quantity_Per_SKU'
            processed_df = processed_df.drop(columns=['Total_Quantity_Per_SKU'])
            
            # Step 6: Remove duplicates in the Product Column (Column A)
            st.write("6. Removing duplicates...")
            processed_df = processed_df.drop_duplicates(subset=[sku_column])
            
            # Step 7: Calculate SKU Velocity
            st.write("7. Calculating SKU Velocity...")
            # Assume Column G is the 7th column (index 6)
            if len(processed_df.columns) > 6:
                column_g_name = processed_df.columns[6]
                processed_df['SKU_Velocity'] = processed_df[column_g_name] / total_quantity * 100
            else:
                st.warning("Column G not found. SKU Velocity calculation skipped.")
            
            # Step 8: Sort by SKU Velocity (largest to smallest)
            st.write("8. Sorting by SKU Velocity...")
            if 'SKU_Velocity' in processed_df.columns:
                processed_df = processed_df.sort_values(by='SKU_Velocity', ascending=False)
            
            # Step 9: Highlight SKUs with 200+ units sold or 80% of sales volume
            st.write("9. Identifying focus SKUs...")
            if 'Total_Quantity_Values' in processed_df.columns and 'SKU_Velocity' in processed_df.columns:
                # Identify SKUs with 200+ units
                high_unit_skus = processed_df[processed_df['Total_Quantity_Values'] >= 200]
                
                # Identify SKUs that make up 80% of sales volume
                processed_df['Cumulative_Percentage'] = processed_df['SKU_Velocity'].cumsum()
                volume_skus = processed_df[processed_df['Cumulative_Percentage'] <= 80]
                
                # Create a column to mark focus SKUs
                processed_df['Focus_SKU'] = False
                processed_df.loc[high_unit_skus.index, 'Focus_SKU'] = True
                processed_df.loc[volume_skus.index, 'Focus_SKU'] = True
            
            # Display the processed data
            st.subheader("Processed Data")
            st.dataframe(processed_df)
            
            # Provide a download link for the processed data
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                processed_df.to_excel(writer, index=False, sheet_name='Cleaned Data')
                
                # Try to access the active worksheet to apply formatting
                try:
                    worksheet = writer.sheets['Cleaned Data']
                    
                    # Apply filter to the header row
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Highlight Focus SKUs
                    if 'Focus_SKU' in processed_df.columns:
                        from openpyxl.styles import PatternFill
                        highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        
                        for idx, row in enumerate(processed_df.itertuples(), start=2):  # Start from row 2 (after header)
                            if row.Focus_SKU:
                                for col in range(1, len(processed_df.columns) + 1):
                                    worksheet.cell(row=idx, column=col).fill = highlight_fill
                except Exception as e:
                    st.warning(f"Could not apply Excel formatting: {e}")
            
            buffer.seek(0)
            
            # Create download button
            b64 = base64.b64encode(buffer.getvalue()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="cleaned_data.xlsx">Download Processed Excel File</a>'
            st.markdown(href, unsafe_allow_html=True)
            
            # Display additional information
            st.subheader("Summary Statistics")
            st.write(f"Total SKUs: {len(processed_df)}")
            st.write(f"Total Quantity: {total_quantity}")
            
            if 'Focus_SKU' in processed_df.columns:
                focus_skus = processed_df[processed_df['Focus_SKU'] == True]
                st.write(f"Focus SKUs (200+ units or 80% of volume): {len(focus_skus)}")
                st.write("Focus SKU List:")
                st.dataframe(focus_skus[[sku_column, 'Total_Quantity_Values', 'SKU_Velocity']])
                
    except Exception as e:
        st.error(f"Error processing the file: {str(e)}")
else:
    st.info("Please upload an Excel file to begin.")

# Add some helpful information at the bottom
st.markdown("---")
st.markdown("### Processing Steps:")
st.markdown("""
1. Calculate total quantity sold per SKU
2. Copy totals as values only
3. Store the total quantity
4. Clear rows past the last product line
5. Remove unnecessary columns
6. Remove duplicates in the Product Column
7. Calculate SKU Velocity (formula: quantity/total * 100)
8. Sort by SKU Velocity (largest to smallest)
9. Identify focus SKUs (200+ units sold or top SKUs comprising 80% of sales)
""")
